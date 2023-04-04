import os
import openpyxl
from openpyxl.styles import PatternFill, Font


class CriaExcel:

    def gera_excel(self):
        financeiro_csv = '/Files/Financeiro.csv'

        # Define o nome do arquivo CSV que será lido
        arquivoCsv = 'Files/Financeiro.csv'

        # Carrega a planilha Excel de base
        planilha = openpyxl.load_workbook(filename='Files/tabela_base.xlsx')

        # Obtém a planilha ativa
        abaPlanilha = planilha.active


        # Abre o arquivo CSV para leitura
        with open(arquivoCsv, 'r', encoding='utf-8') as f:
            # Pula a primeira linha do arquivo CSV
            next(f)

            # Itera sobre cada linha do arquivo CSV e as adiciona na planilha Excel
            linha = 2
            for data in f:
                data = data.strip().split(';')
                # Itera sobre cada célula do arquivo CSV e as adiciona na planilha Excel
                coluna = 1
                for celula in data:
                    numeroDaCelula = openpyxl.utils.cell.get_column_letter(coluna) + str(linha)
                    abaPlanilha[numeroDaCelula] = celula
                    coluna += 1
                linha += 1

        # Define a cor de fundo das células da primeira linha
        fill = PatternFill(start_color='FF8CB6F9', fill_type='solid')
        font = Font(bold=True)
        for coluna in abaPlanilha.iter_cols(min_row=1, max_row=1):
            for celula in coluna:
                celula.fill = fill
                celula.font = font

        # Ajusta a largura das colunas
        tamanhoColuna = [15, 48, 16, 28, 20, 13, 10, 14, 19, 18]
        letraColuna = 0
        for tamanho in tamanhoColuna:
            abaPlanilha.column_dimensions[openpyxl.utils.cell.get_column_letter(letraColuna+1)].width = tamanho
            letraColuna += 1

        # Salva a planilha Excel em um arquivo com nome diferente
        planilha.save(filename='Files/Financeiro.xlsx')
        c = abaPlanilha['K1']
        c.fill = None
        print('Planilha criada com sucesso')
